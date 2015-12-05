#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
 Created on Fri Dec  4 22:27:37 2015

 Author : Philippe Baucour philippe.baucour@univ-fcomte.fr
"""
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,MA 02110-1301,USA.

import openpyxl as xls
import scipy.sparse as scsp
import numpy as np
import matplotlib.pyplot as plt


def read_Excel(filename,sheet,cellrange):
    """ Transform a range of cells in a Excel sheet into a sparse matrix.
    Any string in the range will be converted to the corresponding
    Python variable if it exists or NaN
    
    :param filename: The name of an Excel file
    :param sheet: The sheet title
    :param cellrange: The range of cells in Excel format
    :type filename: str
    :type sheet: str
    :type cellrange: str
    :return: A sparse matrix

    :Example:
    
    * let's have the following Excel file : **Try.xlsx**
    * with a sheet named : **Sheet1**
    * and the following range of cells : **B2:E5**

    +-----+---+---+---+---+
    |     | B | C | D | E |
    +=====+===+===+===+===+
    |**2**| 1 |   |   |   |
    +-----+---+---+---+---+
    |**3**|   | 2 | 3 |   |
    +-----+---+---+---+---+
    |**4**|   | a |   | 1 |
    +-----+---+---+---+---+
    |**5**|   |   | c |   |
    +-----+---+---+---+---+
   

    * and the variable **a** already defined to **-1**
    
    >>> a=-1
    >>> Ms=read_Excel('Try.xlsx','Sheet1','B2:E5')
    Warning: Error in Cell  D5
        c  is not a Python variable
        M[ 3 , 2 ] is set to nan
    
    an error occured at the **Cell D5** for which the **c** variable is 
    not defined in Python. This missing value will be replaced by **NaN**
    
    >>> print Ms
    (0, 0)        1.0
    (1, 1)        2.0
    (1, 2)        3.0
    (2, 1)        -1.0
    (2, 3)        1.0
    (3, 2)        nan
    
    and
    
    >>> print Ms.todense()
    [[  1.   0.   0.   0.]
    [  0.   2.   3.   0.]
    [  0.  -1.   0.   1.]
    [  0.   0.  nan   0.]]
    """
    wb=xls.load_workbook(filename)
    F=wb.get_sheet_by_name(sheet)
    min_col,min_row,max_col,max_row =xls.utils.range_boundaries(cellrange)
    rowindex=range(max_row-min_row+1)
    colindex=range(max_col-min_col+1)
    M=scsp.lil_matrix(np.zeros((len(rowindex),len(colindex))))
    
    for r,idxr in zip(F.iter_rows(range_string=cellrange),rowindex):        
        for cell,idxc in zip(r,colindex) :
            if isinstance(cell.value, (long, int,float)) :
                M[idxr,idxc]=cell.value
            elif cell.value in globals() :
                exec('M[idxr,idxc]='+cell.value)
            elif cell.value<>None :
                posError=xls.utils.get_column_letter(idxc+min_col)+str(idxr+min_row)
                print "Warning: Error in Cell ",posError
                print "\t",cell.value,' is not a Python variable'
                print "\t","M[",idxr,',',idxc,"] is set to nan"
                exec('M[idxr,idxc]=np.nan')

    return M

if __name__ == '__main__':
    plt.close('all')
    a=12.5
    nom='essai.xlsx'
    f='Feuille1'
    selec='B2:E5'
    As=read_Excel(nom,f,selec)
    plt.imshow(As.todense(),interpolation='none')
    plt.colorbar()

